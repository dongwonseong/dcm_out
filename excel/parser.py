import pandas as pd
import os
import re
from util.dbService import DatabaseService
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Color, numbers, Alignment

load_dotenv()
class ExcelProcess:

    def __init__(self, type):
        dbService = DatabaseService()
        self.path = os.getenv('EXCEL_PATH')
        self.data = dbService.execute_query_for_type(type)

        # 증권사 이름 맵핑 예외 추가 될 때 마다 추가
        self.name_mapping = {
            "KB증권": ["케이비증권", "케이비증권(주)", "케이비증권 주식회사", "KB증권", "KB증권(주)"],
            "SK증권": ["SK증권(주)", "에스케이증권", "에스케이증권(주)", "에스케이증권 주식회사", "SK증권"],
            "BNK투자증권": ["BNK투자증권", "BNK투자증권(주)", "BNK투자증권 주식회사"],
            "DB금융투자": ["DB금융투자", "DB금융투자(주)", "DB금융투자 주식회사"],
            "NH투자증권": ["NH투자증권", "NH투자증권(주)", "NH투자증권 주식회사"],
            "교보증권": ["교보증권", "교보증권(주)", "교보증권 주식회사"],
            "다올투자증권": ["다올투자증권", "다올투자증권(주)", "다올투자증권 주식회사"],
            "대신증권": ["대신증권", "대신증권(주)", "대신증권 주식회사"],
            "리딩투자증권": ["리딩투자증권", "리딩투자증권(주)", "리딩투자증권 주식회사"],
            "메리츠증권": ["메리츠증권", "메리츠증권(주)", "메리츠증권 주식회사"],
            "미래에셋증권": ["미래에셋증권", "미래에셋증권(주)", "미래에셋증권 주식회사"],
            "부국증권": ["부국증권", "부국증권(주)", "부국증권 주식회사"],
            "삼성증권": ["삼성증권", "삼성증권(주)", "삼성증권 주식회사"],
            "상상인증권": ["상상인증권", "상상인증권(주)", "상상인증권 주식회사"],
            "신한투자증권": ["신한투자증권", "신한투자증권(주)", "신한투자증권 주식회사"],
            "유안타증권": ["유안타증권", "유안타증권(주)", "유안타증권 주식회사"],
            "유진투자증권": ["유진투자증권", "유진투자증권(주)", "유진투자증권 주식회사", "유진증권"],
            "코리아에셋투자증권": ["코리아에셋투자증권", "코리아에셋투자증권(주)", "코리아에셋투자증권 주식회사"],
            "키움증권": ["키움증권", "키움증권(주)", "키움증권 주식회사"],
            "하나증권": ["하나증권", "하나증권(주)", "하나증권 주식회사"],
            "하이투자증권": ["하이투자증권", "하이투자증권(주)", "하이투자증권 주식회사"],
            "한국산업은행": ["한국산업은행", "한국산업은행(주)", "한국산업은행 주식회사"],
            "한국투자증권": ["한국투자증권", "한국투자증권(주)", "한국투자증권 주식회사"],
            "한양증권": ["한양증권", "한양증권(주)", "한양증권 주식회사"],
            "한화투자증권": ["한화투자증권", "한화투자증권(주)", "한화투자증권 주식회사"],
            "흥국증권": ["흥국증권", "흥국증권(주)", "흥국증권 주식회사"],
            "KIDB채권중개": ["케이아이디비채권중개"],
            "KR투자증권": ["케이알투자증권", "케이알투자증권(주)", "케이알투자증권 주식회사"],
            "IBK투자증권": ["아이비케이투자증권", "아이비케이투자증권(주)", "아이비케이투자증권 주식회사"]
        }


    def excel_process(self):        # 로직 메소드
        self.clean_numeric_data()   # 데이터 정리
        self.standardize_company_names() # 증권사 이름 맵핑
        self.calculate_fee_bp()     # fee_bp (수수료) 계산
        self.calculate_fee_amount() # fee_amount (수수료 정액) 계산
        self.write_to_excel()       # 데이터를 엑셀에 쓰기


    def calculate_fee_bp(self):
        for record in self.data:
            fee_bp = record.get('fee_bp')
            fee_amount = record.get('fee_amount')
            underwriter_amount = record.get('underwriter_amount')

            if not fee_bp and fee_amount and underwriter_amount:
                try:
                    fee_amount_val = float(fee_amount)
                    underwriter_amount_val = float(underwriter_amount)
                    calculated_fee_bp = (fee_amount_val / underwriter_amount_val) * 10000
                    record['fee_bp'] = calculated_fee_bp
                except (ValueError, TypeError):
                    record['fee_bp'] = None

    def calculate_fee_amount(self):
        for record in self.data:
            fee_bp = record.get('fee_bp')
            fee_amount = record.get('fee_amount')
            underwriter_amount = record.get('underwriter_amount')

            if fee_bp and not fee_amount and underwriter_amount:
                try:
                    fee_bp_val = float(fee_bp)
                    underwriter_amount_val = float(underwriter_amount)
                    calculated_value = underwriter_amount_val * fee_bp_val / 10000
                    record['fee_amount'] = calculated_value
                except (ValueError, TypeError):
                    record['fee_amount'] = None

    def write_to_excel(self):
        book = load_workbook(self.path)
        sheet = book.active
        font = Font(name='맑은 고딕', size=8)
        number_format = '#,##0'
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        center_aligned_text = Alignment(horizontal='center')

        issue_date_idx = None
        for idx, cell in enumerate(sheet[1]):
            if cell.value == 'issue_date':
                issue_date_idx = idx + 1
                break

        numeric_columns = ['issue_amount', 'based_issue_amount', 'lead_underwriter_amount',
                           'joint_underwriters_amount', 'underwriter_amount', 'facility_funds',
                           'operating_funds', 'refinancing_funds', 'other_funds', 'fee_amount']

        excluded_columns = ['standard_code', 'corp_code', 'number', 'pv_key', 'method', 'last_update']

        for row_idx, record in enumerate(self.data, start=sheet.max_row + 1):
            method_value = record.get('method')

            col_idx = 1
            for key, value in record.items():
                if key not in excluded_columns:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = font
                    cell.alignment = center_aligned_text

                    if key in numeric_columns and isinstance(value, (int, float)):
                        cell.number_format = number_format

                    if key == 'market_segment' and value == 'X' and issue_date_idx:
                        sheet.cell(row=row_idx, column=issue_date_idx).fill = yellow_fill
                    if key == 'bond_name':
                        value = re.sub(r"\([^)]*\)", "", value)

                    cell.value = value

                    col_idx += 1


        book.save(self.path)

    def clean_numeric_data(self):
        numeric_columns = ['issue_amount', 'based_issue_amount', 'lead_underwriter_amount',
                           'joint_underwriters_amount', 'underwriter_amount', 'facility_funds',
                           'operating_funds', 'refinancing_funds', 'other_funds']
        for record in self.data:
            for col in numeric_columns:
                if col in record and record[col]:
                    try:
                        record[col] = float(record[col].replace(',', ''))
                    except ValueError:
                        print(f"값 변환 오류: {record[col]}")

    def standardize_company_names(self):

        reverse_mapping = {}
        for standard, aliases in self.name_mapping.items():
            for alias in aliases:
                reverse_mapping[alias] = standard

        for record in self.data:
            for key in ['underwriter', 'joint_underwriters', 'lead_underwriter']:
                if key in record:
                    if isinstance(record[key], list):
                        standardized_names = [reverse_mapping.get(name, name) for name in record[key]]
                        record[key] = standardized_names
                    else:
                        record[key] = reverse_mapping.get(record[key], record[key])




